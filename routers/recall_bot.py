"""
Recall Bot Router
Handles Telegram admin panel and scheduled recall reminders.
"""
import asyncio
from datetime import datetime, timedelta, timezone, date
from io import BytesIO
from typing import Optional, Dict
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, ConfigDict
from sqlalchemy import select, and_, insert, update, func
from sqlalchemy.ext.asyncio import AsyncSession
from telegram import Bot, ReplyKeyboardMarkup, KeyboardButton, InputFile

from auth_utils.auth_func import get_current_user
from config import (
    TELEGRAM_RECALL_BOT_TOKEN,
    RECALL_BOT_ADMIN_PASSWORD,
    RECALL_DAILY_STATS_HOUR,
    RECALL_DAILY_STATS_WINDOW_MINUTES,
    RECALL_DAILY_STATS_INTERVAL_DAYS
)
from database import get_async_session, async_session_maker
from models.admin_models import (
    customer,
    CustomerStatus,
    recall_bot_admin,
    recall_bot_recipient,
    recall_notification_log,
    crm_daily_stats_delivery_log,
    customer_status_change_log
)
from models.user_models import user_page_permission, PageName
from utils.crypto import decrypt_text
from utils.ai_summary import generate_customer_ai_summary

router = APIRouter(prefix="/recall-bot", tags=["Recall Bot"])

bot = Bot(token=TELEGRAM_RECALL_BOT_TOKEN) if TELEGRAM_RECALL_BOT_TOKEN else None

REMINDER_MINUTES = (30, 1)
SCHEDULER_INTERVAL_SECONDS = 30
DUE_WINDOW_PAST_SECONDS = 120
DUE_WINDOW_FUTURE_SECONDS = 20
LOOKAHEAD_MINUTES = 61
UZBEKISTAN_TZ = ZoneInfo("Asia/Tashkent")

_scheduler_task: Optional[asyncio.Task] = None

BTN_USERS = "📋 Users"
BTN_RUN = "🚀 Run Reminders"
BTN_ADD = "➕ Add User"
BTN_REMOVE = "➖ Remove User"
BTN_HELP = "❓ Help"
BTN_STATS_MENU = "📈 Statistika"
BTN_STATS_3_DAYS = "🟡 3 kunlik"
BTN_STATS_1_WEEK = "🔵 1 haftalik"
BTN_STATS_1_MONTH = "🟣 1 oylik"
BTN_STATS_EXCEL = "📎 Excel fayl"
BTN_STATS_BACK = "⬅️ Orqaga"
BTN_MY_ID = "🆔 My ID"

STATS_PERIOD_COMMANDS = {
    "/stats_3d": ("last_3_days", "🟡 Oxirgi 3 kun"),
    "/stats_7d": ("last_7_days", "🔵 Oxirgi 1 hafta"),
    "/stats_30d": ("last_30_days", "🟣 Oxirgi 1 oy"),
}
STATUS_KEYS = (
    "need_to_call",
    "contacted",
    "project_started",
    "continuing",
    "finished",
    "rejected",
)


class TelegramUser(BaseModel):
    id: int
    first_name: str
    last_name: Optional[str] = None
    username: Optional[str] = None


class TelegramChat(BaseModel):
    id: int
    type: str
    title: Optional[str] = None


class TelegramMessage(BaseModel):
    message_id: int
    from_: TelegramUser = Field(alias="from")
    chat: TelegramChat
    date: int
    text: Optional[str] = None

    model_config = ConfigDict(populate_by_name=True)


class TelegramWebhookPayload(BaseModel):
    update_id: int
    message: Optional[TelegramMessage] = None
    edited_message: Optional[TelegramMessage] = None


def _normalize_username(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    if value.startswith("@"):
        value = value[1:]
    return value.lower()


def _safe_decrypt(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        return decrypt_text(value)
    except Exception:
        return value


def _utc_now_naive() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _format_uzbek_time_from_utc_naive(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    else:
        value = value.astimezone(timezone.utc)
    return value.astimezone(UZBEKISTAN_TZ).strftime("%Y-%m-%d %H:%M")


def _date_range_uz_to_utc_naive(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    start_uz = datetime(start_date.year, start_date.month, start_date.day, tzinfo=UZBEKISTAN_TZ)
    end_next_uz = datetime(end_date.year, end_date.month, end_date.day, tzinfo=UZBEKISTAN_TZ) + timedelta(days=1)
    return (
        start_uz.astimezone(timezone.utc).replace(tzinfo=None),
        end_next_uz.astimezone(timezone.utc).replace(tzinfo=None)
    )


def _build_status_percentages(status_stats: dict[str, int], total: int) -> dict[str, float]:
    percentages: dict[str, float] = {}
    if total > 0:
        for key, count in status_stats.items():
            percentages[key] = round((count / total) * 100, 1)
    else:
        for key in status_stats.keys():
            percentages[key] = 0.0
    return percentages


def _empty_status_counts() -> dict[str, int]:
    return {key: 0 for key in STATUS_KEYS}


async def _get_new_leads_count_for_range(
    session: AsyncSession,
    start_date: date,
    end_date: date
) -> int:
    start_utc_naive, end_utc_naive = _date_range_uz_to_utc_naive(start_date, end_date)
    result = await session.execute(
        select(func.count(customer.c.id)).where(
            and_(
                customer.c.created_at >= start_utc_naive,
                customer.c.created_at < end_utc_naive
            )
        )
    )
    return int(result.scalar() or 0)


async def _get_status_changes_for_range(
    session: AsyncSession,
    start_date: date,
    end_date: date
) -> dict[str, int]:
    status_changes = _empty_status_counts()
    start_utc_naive, end_utc_naive = _date_range_uz_to_utc_naive(start_date, end_date)
    try:
        result = await session.execute(
            select(
                customer_status_change_log.c.to_status,
                func.count(customer_status_change_log.c.id).label("count")
            ).where(
                and_(
                    customer_status_change_log.c.changed_at >= start_utc_naive,
                    customer_status_change_log.c.changed_at < end_utc_naive
                )
            ).group_by(customer_status_change_log.c.to_status)
        )
        for row in result.fetchall():
            status_key = row.to_status.value if hasattr(row.to_status, "value") else str(row.to_status)
            if status_key in status_changes:
                status_changes[status_key] = int(row.count or 0)
    except Exception:
        return status_changes
    return status_changes


async def _get_crm_status_stats_for_range(
    session: AsyncSession,
    start_date: date,
    end_date: date
) -> dict:
    start_utc_naive, end_utc_naive = _date_range_uz_to_utc_naive(start_date, end_date)
    result = await session.execute(
        select(
            func.count(customer.c.id).label("total_customers"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.need_to_call).label("need_to_call"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.contacted).label("contacted"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.project_started).label("project_started"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.continuing).label("continuing"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.finished).label("finished"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.rejected).label("rejected")
        ).where(
            and_(
                customer.c.created_at >= start_utc_naive,
                customer.c.created_at < end_utc_naive
            )
        )
    )
    row = result.fetchone()
    status_stats = {
        "need_to_call": row.need_to_call,
        "contacted": row.contacted,
        "project_started": row.project_started,
        "continuing": row.continuing,
        "finished": row.finished,
        "rejected": row.rejected
    }


    total = row.total_customers
    return {
        "total_customers": total,
        "status_stats": status_stats,
        "status_percentages": _build_status_percentages(status_stats, total)
    }


async def _get_crm_status_stats_snapshot(session: AsyncSession) -> dict:
    result = await session.execute(
        select(
            func.count(customer.c.id).label("total_customers"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.need_to_call).label("need_to_call"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.contacted).label("contacted"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.project_started).label("project_started"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.continuing).label("continuing"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.finished).label("finished"),
            func.count(customer.c.id).filter(customer.c.status == CustomerStatus.rejected).label("rejected")
        )
    )
    row = result.fetchone()
    status_stats = {
        "need_to_call": row.need_to_call,
        "contacted": row.contacted,
        "project_started": row.project_started,
        "continuing": row.continuing,
        "finished": row.finished,
        "rejected": row.rejected
    }
    total = row.total_customers
    return {
        "total_customers": total,
        "status_stats": status_stats,
        "status_percentages": _build_status_percentages(status_stats, total)
    }


def _format_period_block(title: str, data: dict) -> str:
    status = data["status_stats"]
    return (
        f"{title}\n"
        f"Jami: {data['total_customers']}\n"
        f"need_to_call: {status['need_to_call']}\n"
        f"contacted: {status['contacted']}\n"
        f"project_started: {status['project_started']}\n"
        f"continuing: {status['continuing']}\n"
        f"finished: {status['finished']}\n"
        f"rejected: {status['rejected']}"
    )


def _clean_note_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    cleaned = " ".join(str(value).split()).strip()
    return cleaned or None


async def _collect_notes_for_range(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    limit: int = 80
) -> list[str]:
    start_utc_naive, end_utc_naive = _date_range_uz_to_utc_naive(start_date, end_date)
    result = await session.execute(
        select(
            customer.c.notes,
            customer.c.status
        ).where(
            and_(
                customer.c.notes.isnot(None),
                customer.c.created_at >= start_utc_naive,
                customer.c.created_at < end_utc_naive
            )
        ).order_by(customer.c.created_at.desc()).limit(limit)
    )

    notes: list[str] = []
    for row in result.fetchall():
        cleaned = _clean_note_text(row.notes)
        if not cleaned:
            continue
        status_name = row.status.value if hasattr(row.status, "value") else str(row.status)
        notes.append(f"[{status_name}] {cleaned}")
    return notes


async def _build_ai_notes_summary_block(
    session: AsyncSession,
    report_date: date
) -> Optional[str]:
    notes = await _collect_notes_for_range(session, report_date, report_date)
    source_label = "bugungi notelar"

    if not notes:
        notes = await _collect_notes_for_range(session, report_date - timedelta(days=2), report_date)
        source_label = "oxirgi 3 kun notelari"

    if not notes:
        return "AI xulosa: bugun va oxirgi 3 kunda notes topilmadi."

    notes_payload = "\n".join(notes)
    if len(notes_payload) > 7000:
        notes_payload = notes_payload[:7000]

    ai_summary = await generate_customer_ai_summary(notes_payload)
    if not ai_summary:
        return f"AI xulosa ({source_label}): qisqa xulosa shakllantirib bo'lmadi."

    return f"AI xulosa ({source_label}):\n{ai_summary}"


async def _collect_daily_crm_stats_payload(
    session: AsyncSession,
    report_date: date,
    include_ai_summary: bool = True,
    use_snapshot_statuses: bool = False
) -> dict:
    period_ranges = {
        "today": (report_date, report_date),
        "last_3_days": (report_date - timedelta(days=2), report_date),
        "last_7_days": (report_date - timedelta(days=6), report_date),
        "last_30_days": (report_date - timedelta(days=29), report_date),
        "last_90_days": (report_date - timedelta(days=89), report_date),
    }

    period_payloads: dict[str, dict] = {}
    if use_snapshot_statuses:
        snapshot = await _get_crm_status_stats_snapshot(session)
        for period_key, (start_date, end_date) in period_ranges.items():
            period_data = {
                "total_customers": snapshot["total_customers"],
                "status_stats": dict(snapshot["status_stats"]),
                "status_percentages": dict(snapshot["status_percentages"]),
            }
            period_data["new_leads_count"] = await _get_new_leads_count_for_range(session, start_date, end_date)
            period_data["status_changes"] = await _get_status_changes_for_range(session, start_date, end_date)
            period_payloads[period_key] = period_data
    else:
        for period_key, (start_date, end_date) in period_ranges.items():
            period_data = await _get_crm_status_stats_for_range(session, start_date, end_date)
            period_data["new_leads_count"] = await _get_new_leads_count_for_range(session, start_date, end_date)
            period_data["status_changes"] = await _get_status_changes_for_range(session, start_date, end_date)
            period_payloads[period_key] = period_data

    ai_notes_summary = None
    if include_ai_summary:
        ai_notes_summary = await _build_ai_notes_summary_block(session, report_date)

    return {
        "today": period_payloads["today"],
        "last_3_days": period_payloads["last_3_days"],
        "last_7_days": period_payloads["last_7_days"],
        "last_30_days": period_payloads["last_30_days"],
        "last_90_days": period_payloads["last_90_days"],
        "ai_notes_summary": ai_notes_summary
    }


def _build_daily_crm_stats_text_from_payload(report_date: date, payload: dict) -> str:
    today = payload["today"]
    last_3_days = payload["last_3_days"]
    last_7_days = payload["last_7_days"]
    last_30_days = payload["last_30_days"]
    last_90_days = payload["last_90_days"]
    ai_notes_summary = payload.get("ai_notes_summary")
    text = (
        f"📊 CRM Kunlik Statistika\n"
        f"📅 Sana: {report_date.isoformat()} (UZ)\n\n"
        f"{_format_period_block('🟢 Bugun', today)}\n\n"
        f"{_format_period_block('🟡 Oxirgi 3 kun', last_3_days)}\n\n"
        f"{_format_period_block('🔵 Oxirgi 1 hafta', last_7_days)}\n\n"
        f"{_format_period_block('🟣 Oxirgi 1 oy', last_30_days)}\n\n"
        f"{_format_period_block('🟤 Oxirgi 3 oy', last_90_days)}"
    )
    if ai_notes_summary:
        text += f"\n\n{ai_notes_summary}"
    return text


def _build_period_stats_text_from_payload(
    report_date: date,
    payload: dict,
    period_key: str,
    period_title: str
) -> str:
    period_payload = payload.get(period_key)
    if not period_payload:
        return (
            "📊 CRM Statistika\n"
            f"📅 Sana: {report_date.isoformat()} (UZ)\n\n"
            "⚠️ Tanlangan davr uchun ma'lumot topilmadi."
        )
    status_changes = period_payload.get("status_changes") or _empty_status_counts()
    new_leads_count = int(period_payload.get("new_leads_count") or 0)
    status_changes_text = (
        f"need_to_call: {status_changes['need_to_call']}\n"
        f"contacted: {status_changes['contacted']}\n"
        f"project_started: {status_changes['project_started']}\n"
        f"continuing: {status_changes['continuing']}\n"
        f"finished: {status_changes['finished']}\n"
        f"rejected: {status_changes['rejected']}"
    )

    return (
        "📊 CRM Statistika\n"
        f"📅 Sana: {report_date.isoformat()} (UZ)\n\n"
        f"{_format_period_block(f'{period_title} (joriy statuslar)', period_payload)}\n\n"
        f"🆕 Kelgan leadlar: {new_leads_count}\n\n"
        "🔄 Status o'zgarganlar (davr ichida):\n"
        f"{status_changes_text}"
    )


def _build_daily_crm_stats_excel(report_date: date, payload: dict) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CRM Stats"

    period_rows = [
        ("🟢 Bugun", payload["today"], "D9EAD3"),
        ("🟡 Oxirgi 3 kun", payload["last_3_days"], "FFF2CC"),
        ("🔵 Oxirgi 1 hafta", payload["last_7_days"], "CFE2F3"),
        ("🟣 Oxirgi 1 oy", payload["last_30_days"], "EAD1DC"),
        ("🟤 Oxirgi 3 oy", payload["last_90_days"], "D9D2E9"),
    ]
    status_order = [
        "need_to_call",
        "contacted",
        "project_started",
        "continuing",
        "finished",
        "rejected",
    ]
    status_labels = {
        "need_to_call": "Need To Call",
        "contacted": "Contacted",
        "project_started": "Project Started",
        "continuing": "Continuing",
        "finished": "Finished",
        "rejected": "Rejected",
    }
    status_fill_map = {
        "need_to_call": "DDEBF7",
        "contacted": "D9EAD3",
        "project_started": "CFE2F3",
        "continuing": "FFF2CC",
        "finished": "E2F0D9",
        "rejected": "F4CCCC",
    }

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    sub_header_fill = PatternFill(start_color="274E13", end_color="274E13", fill_type="solid")
    total_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin_side = Side(style="thin", color="D0D0D0")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    sheet.merge_cells("A1:E1")
    title_cell = sheet["A1"]
    title_cell.value = "CRM Daily Statistics Report"
    title_cell.fill = header_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = center

    sheet.merge_cells("A2:E2")
    date_cell = sheet["A2"]
    date_cell.value = f"Sana (UZ): {report_date.isoformat()}"
    date_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    date_cell.font = bold_font
    date_cell.alignment = center

    row = 4
    for title, data, period_color in period_rows:
        status_stats = data["status_stats"]
        percentages = data["status_percentages"]

        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        period_cell = sheet.cell(row=row, column=1, value=title)
        period_cell.fill = PatternFill(start_color=period_color, end_color=period_color, fill_type="solid")
        period_cell.font = Font(bold=True, size=12)
        period_cell.alignment = left_wrap
        row += 1

        headers = ["Status", "Count", "Percent", "Date", "Period"]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx, value=header)
            cell.fill = sub_header_fill
            cell.font = white_font
            cell.alignment = center
            cell.border = border
        row += 1

        for status_key in status_order:
            status_cell = sheet.cell(row=row, column=1, value=status_labels[status_key])
            count_cell = sheet.cell(row=row, column=2, value=status_stats.get(status_key, 0))
            percent_value = float(percentages.get(status_key, 0.0))
            percent_cell = sheet.cell(row=row, column=3, value=percent_value / 100)
            percent_cell.number_format = "0.0%"
            date_value_cell = sheet.cell(row=row, column=4, value=report_date.isoformat())
            period_value_cell = sheet.cell(row=row, column=5, value=title)

            row_fill = PatternFill(
                start_color=status_fill_map.get(status_key, "FFFFFF"),
                end_color=status_fill_map.get(status_key, "FFFFFF"),
                fill_type="solid"
            )
            for cell in (status_cell, count_cell, percent_cell, date_value_cell, period_value_cell):
                cell.fill = row_fill
                cell.border = border
                cell.alignment = center
            status_cell.alignment = left_wrap
            row += 1

        for col_idx in range(1, 6):
            total_cell = sheet.cell(row=row, column=col_idx)
            total_cell.fill = total_fill
            total_cell.border = border
            total_cell.alignment = center
            total_cell.font = bold_font
        sheet.cell(row=row, column=1, value="Jami")
        sheet.cell(row=row, column=2, value=data["total_customers"])
        sheet.cell(row=row, column=3, value=1.0 if data["total_customers"] else 0.0).number_format = "0.0%"
        row += 2

    ai_summary = payload.get("ai_notes_summary")
    if ai_summary:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ai_title = sheet.cell(row=row, column=1, value="AI Xulosa")
        ai_title.fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        ai_title.font = bold_font
        ai_title.alignment = left_wrap
        row += 1

        sheet.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=5)
        ai_text = sheet.cell(row=row, column=1, value=ai_summary)
        ai_text.alignment = left_wrap
        ai_text.border = border

    column_widths = {
        1: 24,
        2: 12,
        3: 12,
        4: 16,
        5: 22,
    }
    for col_idx, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


async def _build_daily_crm_stats_text(session: AsyncSession, report_date: date) -> str:
    payload = await _collect_daily_crm_stats_payload(session, report_date)
    return _build_daily_crm_stats_text_from_payload(report_date, payload)


def _admin_keyboard() -> ReplyKeyboardMarkup:
    keyboard = [
        [KeyboardButton(BTN_USERS)],
        [KeyboardButton(BTN_STATS_MENU)],
        [KeyboardButton(BTN_ADD), KeyboardButton(BTN_REMOVE)],
        [KeyboardButton(BTN_HELP)]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=False
    )


def _user_keyboard() -> ReplyKeyboardMarkup:
    keyboard = [
        [KeyboardButton(BTN_STATS_MENU)],
        [KeyboardButton(BTN_HELP), KeyboardButton(BTN_MY_ID)]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=False
    )


def _stats_keyboard() -> ReplyKeyboardMarkup:
    keyboard = [
        [KeyboardButton(BTN_STATS_3_DAYS), KeyboardButton(BTN_STATS_1_WEEK)],
        [KeyboardButton(BTN_STATS_1_MONTH), KeyboardButton(BTN_STATS_EXCEL)],
        [KeyboardButton(BTN_STATS_BACK)]
    ]
    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True,
        one_time_keyboard=False
    )


def _admin_help_text() -> str:
    return (
        "🎛 Recall Bot Admin Panel\n\n"
        "📌 /panel - admin panelni ochish\n"
        "📈 /stats_menu - individual statistika menyusi\n"
        "📋 /list_users - xabar oluvchilar ro'yxati\n"
        "➕ /add_user <chat_id> [username] - yangi user qo'shish\n"
        "➖ /remove_user <chat_id> - userni o'chirish\n"
        "🚀 /run - hozirning o'zida reminderlarni tekshirish\n"
        "🆔 /myid - o'zingizning chat_id ni ko'rish"
    )


def _basic_help_text(chat_id: str) -> str:
    return (
        "🤖 Recall Botga xush kelibsiz.\n"
        f"🆔 Sizning chat_id: {chat_id}\n\n"
        "📈 Statistika olish uchun: /stats_menu\n"
        "🔐 Admin panel uchun: /admin <parol>\n"
        "🆔 Chat ID olish uchun: /myid"
    )


def _stats_menu_text() -> str:
    return (
        "📈 Statistika menyusi\n\n"
        "Kerakli tugmani bosing:\n"
        f"- {BTN_STATS_3_DAYS}\n"
        f"- {BTN_STATS_1_WEEK}\n"
        f"- {BTN_STATS_1_MONTH}\n"
        f"- {BTN_STATS_EXCEL}"
    )


def _button_to_command(text: str) -> str:
    mapping = {
        BTN_USERS: "/list_users",
        BTN_RUN: "/run",
        BTN_ADD: "/add_user",
        BTN_REMOVE: "/remove_user",
        BTN_HELP: "/help",
        BTN_STATS_MENU: "/stats_menu",
        BTN_STATS_3_DAYS: "/stats_3d",
        BTN_STATS_1_WEEK: "/stats_7d",
        BTN_STATS_1_MONTH: "/stats_30d",
        BTN_STATS_EXCEL: "/stats_excel",
        BTN_STATS_BACK: "/stats_back",
        BTN_MY_ID: "/myid",
        "Users": "/list_users",
        "Run Reminders": "/run",
        "Add User": "/add_user",
        "Remove User": "/remove_user",
        "Help": "/help",
        "Statistika": "/stats_menu",
        "3 kunlik": "/stats_3d",
        "1 haftalik": "/stats_7d",
        "1 oylik": "/stats_30d",
        "Excel fayl": "/stats_excel",
        "Orqaga": "/stats_back",
        "My ID": "/myid",
    }
    return mapping.get(text, text)


async def _send_message(chat_id: str, text: str, reply_markup=None) -> None:
    if not bot:
        raise RuntimeError("TELEGRAM_RECALL_BOT_TOKEN sozlanmagan")
    chat_target = chat_id
    if isinstance(chat_id, str) and chat_id.lstrip("-").isdigit():
        chat_target = int(chat_id)
    await bot.send_message(
        chat_id=chat_target,
        text=text,
        reply_markup=reply_markup
    )


async def _send_document(chat_id: str, file_bytes: bytes, filename: str, caption: Optional[str] = None) -> None:
    if not bot:
        raise RuntimeError("TELEGRAM_RECALL_BOT_TOKEN sozlanmagan")
    chat_target = chat_id
    if isinstance(chat_id, str) and chat_id.lstrip("-").isdigit():
        chat_target = int(chat_id)
    await bot.send_document(
        chat_id=chat_target,
        document=InputFile(BytesIO(file_bytes), filename=filename),
        caption=caption
    )


async def _is_admin(session: AsyncSession, chat_id: str) -> bool:
    result = await session.execute(
        select(recall_bot_admin.c.id).where(
            and_(
                recall_bot_admin.c.chat_id == chat_id,
                recall_bot_admin.c.is_active == True
            )
        )
    )
    return result.fetchone() is not None


async def _is_recipient(session: AsyncSession, chat_id: str) -> bool:
    result = await session.execute(
        select(recall_bot_recipient.c.id).where(
            and_(
                recall_bot_recipient.c.chat_id == chat_id,
                recall_bot_recipient.c.is_active == True
            )
        )
    )
    return result.fetchone() is not None


async def _upsert_admin(
    session: AsyncSession,
    chat_id: str,
    telegram_username: Optional[str],
    full_name: Optional[str]
) -> None:
    existing = await session.execute(
        select(recall_bot_admin).where(recall_bot_admin.c.chat_id == chat_id)
    )
    existing_row = existing.fetchone()

    if existing_row:
        await session.execute(
            update(recall_bot_admin)
            .where(recall_bot_admin.c.chat_id == chat_id)
            .values(
                telegram_username=telegram_username,
                full_name=full_name,
                is_active=True,
                updated_at=_utc_now_naive()
            )
        )
    else:
        await session.execute(
            insert(recall_bot_admin).values(
                chat_id=chat_id,
                telegram_username=telegram_username,
                full_name=full_name,
                is_active=True,
                created_at=_utc_now_naive(),
                updated_at=_utc_now_naive()
            )
        )


async def _upsert_recipient(
    session: AsyncSession,
    chat_id: str,
    telegram_username: Optional[str],
    full_name: Optional[str],
    added_by_chat_id: str
) -> None:
    existing = await session.execute(
        select(recall_bot_recipient).where(recall_bot_recipient.c.chat_id == chat_id)
    )
    existing_row = existing.fetchone()

    if existing_row:
        await session.execute(
            update(recall_bot_recipient)
            .where(recall_bot_recipient.c.chat_id == chat_id)
            .values(
                telegram_username=telegram_username,
                full_name=full_name,
                added_by_chat_id=added_by_chat_id,
                is_active=True,
                updated_at=_utc_now_naive()
            )
        )
    else:
        await session.execute(
            insert(recall_bot_recipient).values(
                chat_id=chat_id,
                telegram_username=telegram_username,
                full_name=full_name,
                added_by_chat_id=added_by_chat_id,
                is_active=True,
                created_at=_utc_now_naive(),
                updated_at=_utc_now_naive()
            )
        )


async def _remove_recipient(session: AsyncSession, chat_id: str) -> bool:
    result = await session.execute(
        select(recall_bot_recipient.c.id).where(
            and_(
                recall_bot_recipient.c.chat_id == chat_id,
                recall_bot_recipient.c.is_active == True
            )
        )
    )
    row = result.fetchone()
    if not row:
        return False

    await session.execute(
        update(recall_bot_recipient)
        .where(recall_bot_recipient.c.chat_id == chat_id)
        .values(is_active=False, updated_at=_utc_now_naive())
    )
    return True


async def process_due_recall_notifications(session: AsyncSession) -> Dict[str, int]:
    stats = {
        "customers_checked": 0,
        "recipients_count": 0,
        "due_notifications": 0,
        "sent": 0,
        "failed": 0,
        "skipped_duplicate": 0
    }

    if not bot:
        return stats

    now = _utc_now_naive()
    candidate_start = now - timedelta(minutes=1)
    candidate_end = now + timedelta(minutes=LOOKAHEAD_MINUTES)

    recipients_result = await session.execute(
        select(
            recall_bot_recipient.c.chat_id,
            recall_bot_recipient.c.telegram_username
        ).where(recall_bot_recipient.c.is_active == True)
    )
    recipients = recipients_result.fetchall()
    stats["recipients_count"] = len(recipients)
    if not recipients:
        return stats

    customers_result = await session.execute(
        select(
            customer.c.id,
            customer.c.full_name,
            customer.c.phone_number,
            customer.c.notes,
            customer.c.recall_time
        ).where(
            and_(
                customer.c.recall_time.isnot(None),
                customer.c.recall_time >= candidate_start,
                customer.c.recall_time <= candidate_end
            )
        )
    )
    customers = customers_result.fetchall()
    stats["customers_checked"] = len(customers)
    if not customers:
        return stats

    existing_logs_result = await session.execute(
        select(
            recall_notification_log.c.customer_id,
            recall_notification_log.c.recipient_chat_id,
            recall_notification_log.c.reminder_minutes,
            recall_notification_log.c.scheduled_for
        ).where(
            and_(
                recall_notification_log.c.scheduled_for >= candidate_start,
                recall_notification_log.c.scheduled_for <= candidate_end
            )
        )
    )
    existing_keys = {
        (r.customer_id, r.recipient_chat_id, r.reminder_minutes, r.scheduled_for)
        for r in existing_logs_result.fetchall()
    }

    due_window_start = now - timedelta(seconds=DUE_WINDOW_PAST_SECONDS)
    due_window_end = now + timedelta(seconds=DUE_WINDOW_FUTURE_SECONDS)

    for c in customers:
        recall_at = c.recall_time
        if not recall_at:
            continue

        customer_name = _safe_decrypt(c.full_name)
        customer_phone = _safe_decrypt(c.phone_number)
        customer_note = _clean_note_text(c.notes) or "yo'q"

        for reminder_minutes in REMINDER_MINUTES:
            reminder_at = recall_at - timedelta(minutes=reminder_minutes)
            if not (due_window_start <= reminder_at <= due_window_end):
                continue

            for recipient in recipients:
                key = (c.id, recipient.chat_id, reminder_minutes, recall_at)
                if key in existing_keys:
                    stats["skipped_duplicate"] += 1
                    continue

                stats["due_notifications"] += 1
                insert_result = await session.execute(
                    insert(recall_notification_log).values(
                        customer_id=c.id,
                        recipient_chat_id=recipient.chat_id,
                        reminder_minutes=reminder_minutes,
                        scheduled_for=recall_at,
                        notification_sent=False,
                        created_at=_utc_now_naive()
                    )
                )
                log_id = insert_result.inserted_primary_key[0]
                existing_keys.add(key)

                message_text = (
                    f"🔔 Recall eslatma: {reminder_minutes} daqiqa qoldi\n\n"
                    f"👤 Mijoz: {customer_name}\n"
                    f"📞 Telefon: {customer_phone}\n"
                    f"🕒 Recall vaqti: {_format_uzbek_time_from_utc_naive(recall_at)}\n"
                    f"📝 Note: {customer_note}\n\n"
                    f"✅ Shu customer bilan bog'lanishingiz kerak."
                )

                try:
                    await _send_message(recipient.chat_id, message_text)
                    await session.execute(
                        update(recall_notification_log)
                        .where(recall_notification_log.c.id == log_id)
                        .values(
                            notification_sent=True,
                            sent_at=_utc_now_naive(),
                            error_message=None
                        )
                    )
                    stats["sent"] += 1
                except Exception as exc:
                    await session.execute(
                        update(recall_notification_log)
                        .where(recall_notification_log.c.id == log_id)
                        .values(
                            notification_sent=False,
                            sent_at=None,
                            error_message=str(exc)[:500]
                        )
                    )
                    stats["failed"] += 1

    await session.commit()
    return stats


async def process_daily_crm_stats_digest(
    session: AsyncSession,
    force: bool = False
) -> Dict[str, int]:
    stats = {
        "recipients_count": 0,
        "sent": 0,
        "failed": 0,
        "skipped_already_sent": 0,
        "skipped_time_window": 0,
        "skipped_interval": 0
    }

    if not bot:
        return stats

    now_uz = datetime.now(UZBEKISTAN_TZ)
    report_date = now_uz.date()

    in_window = (
        now_uz.hour == RECALL_DAILY_STATS_HOUR and
        now_uz.minute < RECALL_DAILY_STATS_WINDOW_MINUTES
    )
    if not force and not in_window:
        stats["skipped_time_window"] = 1
        return stats

    if not force:
        has_today_logs_result = await session.execute(
            select(func.count())
            .select_from(crm_daily_stats_delivery_log)
            .where(crm_daily_stats_delivery_log.c.report_date == report_date)
        )
        has_today_logs = bool(has_today_logs_result.scalar() or 0)
        if not has_today_logs:
            last_sent_date_result = await session.execute(
                select(func.max(crm_daily_stats_delivery_log.c.report_date)).where(
                    crm_daily_stats_delivery_log.c.notification_sent == True
                )
            )
            last_sent_date = last_sent_date_result.scalar()
            if (
                last_sent_date and
                (report_date - last_sent_date).days < RECALL_DAILY_STATS_INTERVAL_DAYS
            ):
                stats["skipped_interval"] = 1
                return stats

    recipients_result = await session.execute(
        select(
            recall_bot_recipient.c.chat_id,
            recall_bot_recipient.c.telegram_username
        ).where(recall_bot_recipient.c.is_active == True)
    )
    recipients = recipients_result.fetchall()
    stats["recipients_count"] = len(recipients)
    if not recipients:
        return stats

    existing_result = await session.execute(
        select(
            crm_daily_stats_delivery_log.c.recipient_chat_id,
            crm_daily_stats_delivery_log.c.notification_sent,
            crm_daily_stats_delivery_log.c.id
        ).where(crm_daily_stats_delivery_log.c.report_date == report_date)
    )
    existing_map = {row.recipient_chat_id: row for row in existing_result.fetchall()}

    digest_payload = await _collect_daily_crm_stats_payload(
        session,
        report_date,
        include_ai_summary=False
    )
    digest_text = _build_period_stats_text_from_payload(
        report_date,
        digest_payload,
        "last_3_days",
        "🟡 Oxirgi 3 kun"
    )

    for recipient in recipients:
        existing = existing_map.get(recipient.chat_id)
        if existing and existing.notification_sent and not force:
            stats["skipped_already_sent"] += 1
            continue

        try:
            await _send_message(recipient.chat_id, digest_text)

            if existing:
                await session.execute(
                    update(crm_daily_stats_delivery_log)
                    .where(crm_daily_stats_delivery_log.c.id == existing.id)
                    .values(
                        notification_sent=True,
                        sent_at=_utc_now_naive(),
                        error_message=None
                    )
                )
            else:
                await session.execute(
                    insert(crm_daily_stats_delivery_log).values(
                        report_date=report_date,
                        recipient_chat_id=recipient.chat_id,
                        notification_sent=True,
                        sent_at=_utc_now_naive(),
                        created_at=_utc_now_naive()
                    )
                )
            stats["sent"] += 1
        except Exception as exc:
            if existing:
                await session.execute(
                    update(crm_daily_stats_delivery_log)
                    .where(crm_daily_stats_delivery_log.c.id == existing.id)
                    .values(
                        notification_sent=False,
                        sent_at=None,
                        error_message=str(exc)[:500]
                    )
                )
            else:
                await session.execute(
                    insert(crm_daily_stats_delivery_log).values(
                        report_date=report_date,
                        recipient_chat_id=recipient.chat_id,
                        notification_sent=False,
                        error_message=str(exc)[:500],
                        created_at=_utc_now_naive()
                    )
                )
            stats["failed"] += 1

    await session.commit()
    return stats


async def _scheduler_loop() -> None:
    while True:
        try:
            if bot:
                async with async_session_maker() as session:
                    await process_due_recall_notifications(session)
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            print(f"[recall-bot] scheduler error: {exc}")
        await asyncio.sleep(SCHEDULER_INTERVAL_SECONDS)


@router.on_event("startup")
async def start_recall_scheduler() -> None:
    global _scheduler_task
    if _scheduler_task is None or _scheduler_task.done():
        _scheduler_task = asyncio.create_task(_scheduler_loop())


@router.on_event("shutdown")
async def stop_recall_scheduler() -> None:
    global _scheduler_task
    if _scheduler_task and not _scheduler_task.done():
        _scheduler_task.cancel()
        try:
            await _scheduler_task
        except asyncio.CancelledError:
            pass


@router.get("/recipients", summary="Recall bot recipientlar ro'yxati")
async def get_recall_recipients(
    session: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user)
):
    permissions_result = await session.execute(
        select(user_page_permission.c.page_name).where(
            and_(
                user_page_permission.c.user_id == current_user.id,
                user_page_permission.c.page_name == PageName.crm
            )
        )
    )
    if not permissions_result.fetchone() and current_user.company_code != "ceo":
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    result = await session.execute(
        select(
            recall_bot_recipient.c.chat_id,
            recall_bot_recipient.c.telegram_username,
            recall_bot_recipient.c.full_name,
            recall_bot_recipient.c.is_active
        ).order_by(recall_bot_recipient.c.id.desc())
    )

    return {"items": [dict(r._mapping) for r in result.fetchall()]}


@router.post("/process-reminders", summary="Recall reminderlarni majburan ishga tushirish")
async def run_recall_reminders(
    session: AsyncSession = Depends(get_async_session),
    current_user=Depends(get_current_user)
):
    permissions_result = await session.execute(
        select(user_page_permission.c.page_name).where(
            and_(
                user_page_permission.c.user_id == current_user.id,
                user_page_permission.c.page_name == PageName.crm
            )
        )
    )
    if not permissions_result.fetchone() and current_user.company_code != "ceo":
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    stats = await process_due_recall_notifications(session)
    return {"status": "ok", "stats": stats}


async def _handle_command(
    message: TelegramMessage,
    session: AsyncSession
) -> Dict:
    chat_id = str(message.chat.id)
    username = _normalize_username(message.from_.username)
    full_name = f"{message.from_.first_name} {message.from_.last_name or ''}".strip()

    raw_text = (message.text or "").strip()
    text = _button_to_command(raw_text)
    parts = text.split()
    if not parts:
        return {"status": "ignored", "reason": "empty_message"}

    command = parts[0].lower()

    if command == "/start":
        is_admin = await _is_admin(session, chat_id)
        reply_markup = _admin_keyboard() if is_admin else _user_keyboard()
        welcome_text = _admin_help_text() if is_admin else _basic_help_text(chat_id)
        await _send_message(
            chat_id=chat_id,
            text=welcome_text,
            reply_markup=reply_markup
        )
        return {"status": "success", "command": command}

    if command == "/help":
        is_admin = await _is_admin(session, chat_id)
        if is_admin:
            await _send_message(
                chat_id=chat_id,
                text=_admin_help_text(),
                reply_markup=_admin_keyboard()
            )
        else:
            await _send_message(
                chat_id=chat_id,
                text=_basic_help_text(chat_id),
                reply_markup=_user_keyboard()
            )
        return {"status": "success", "command": command}

    if command == "/myid":
        is_admin = await _is_admin(session, chat_id)
        reply_markup = _admin_keyboard() if is_admin else _user_keyboard()
        await _send_message(chat_id=chat_id, text=f"🆔 Sizning chat_id: {chat_id}", reply_markup=reply_markup)
        return {"status": "success", "command": command}

    if command == "/admin":
        if len(parts) < 2:
            await _send_message(chat_id=chat_id, text="🔐 Parol kiriting: /admin <parol>")
            return {"status": "waiting_password"}

        if parts[1] != RECALL_BOT_ADMIN_PASSWORD:
            await _send_message(chat_id=chat_id, text="❌ Noto'g'ri parol.")
            return {"status": "error", "reason": "wrong_password"}

        await _upsert_admin(
            session=session,
            chat_id=chat_id,
            telegram_username=username,
            full_name=full_name
        )
        await session.commit()

        await _send_message(
            chat_id=chat_id,
            text="✅ Admin panel faollashtirildi.\n\n" + _admin_help_text(),
            reply_markup=_admin_keyboard()
        )
        return {"status": "success", "command": command}

    stats_commands = {"/stats_menu", "/stats_back", "/stats_excel", *STATS_PERIOD_COMMANDS.keys()}
    if command in stats_commands:
        is_admin = await _is_admin(session, chat_id)
        is_recipient = await _is_recipient(session, chat_id)
        if not (is_admin or is_recipient):
            await _send_message(
                chat_id=chat_id,
                text=(
                    "⛔ Statistika ko'rish uchun ruxsat yo'q.\n"
                    "Admin sizni avval /add_user orqali recipientga qo'shishi kerak."
                ),
                reply_markup=_user_keyboard()
            )
            return {"status": "error", "reason": "not_allowed_stats"}

        if command == "/stats_back":
            if is_admin:
                await _send_message(chat_id=chat_id, text=_admin_help_text(), reply_markup=_admin_keyboard())
            else:
                await _send_message(chat_id=chat_id, text=_basic_help_text(chat_id), reply_markup=_user_keyboard())
            return {"status": "success", "command": command}

        if command == "/stats_menu":
            await _send_message(
                chat_id=chat_id,
                text=_stats_menu_text(),
                reply_markup=_stats_keyboard()
            )
            return {"status": "success", "command": command}

        now_uz = datetime.now(UZBEKISTAN_TZ)
        report_date = now_uz.date()

        if command == "/stats_excel":
            payload = await _collect_daily_crm_stats_payload(
                session,
                report_date,
                include_ai_summary=True,
                use_snapshot_statuses=True
            )
            excel_bytes = _build_daily_crm_stats_excel(report_date, payload)
            filename = f"crm_stats_full_{report_date.isoformat()}.xlsx"
            await _send_document(
                chat_id=chat_id,
                file_bytes=excel_bytes,
                filename=filename,
                caption="📎 CRM to'liq statistika (Excel)"
            )
            return {"status": "success", "command": command}

        period_key, period_title = STATS_PERIOD_COMMANDS[command]
        payload = await _collect_daily_crm_stats_payload(
            session,
            report_date,
            include_ai_summary=False,
            use_snapshot_statuses=True
        )
        text = _build_period_stats_text_from_payload(report_date, payload, period_key, period_title)
        await _send_message(chat_id=chat_id, text=text, reply_markup=_stats_keyboard())
        return {"status": "success", "command": command}

    is_admin = await _is_admin(session, chat_id)
    if not is_admin:
        await _send_message(
            chat_id=chat_id,
            text="⛔ Admin emas. /admin <parol> yuboring.",
            reply_markup=_user_keyboard()
        )
        return {"status": "error", "reason": "not_admin"}

    if command == "/panel":
        await _send_message(
            chat_id=chat_id,
            text=_admin_help_text(),
            reply_markup=_admin_keyboard()
        )
        return {"status": "success", "command": command}

    if command == "/list_users":
        result = await session.execute(
            select(
                recall_bot_recipient.c.chat_id,
                recall_bot_recipient.c.telegram_username,
                recall_bot_recipient.c.full_name
            ).where(recall_bot_recipient.c.is_active == True)
            .order_by(recall_bot_recipient.c.id.desc())
        )
        rows = result.fetchall()
        if not rows:
            await _send_message(chat_id=chat_id, text="📭 Active userlar yo'q.")
        else:
            lines = ["📋 Active userlar:"]
            for i, row in enumerate(rows, 1):
                uname = f"@{row.telegram_username}" if row.telegram_username else "username yo'q"
                lines.append(f"{i}. 👤 {row.chat_id} ({uname})")
            await _send_message(chat_id=chat_id, text="\n".join(lines))
        return {"status": "success", "command": command}

    if command == "/add_user":
        if len(parts) < 2:
            await _send_message(
                chat_id=chat_id,
                text="ℹ️ Format: /add_user <chat_id> [username]\nMisol: /add_user 123456789 johndoe"
            )
            return {"status": "error", "reason": "invalid_arguments"}

        target_chat_id = parts[1].strip()
        if target_chat_id.startswith("@"):
            await _send_message(
                chat_id=chat_id,
                text="⚠️ Username bilan yuborish ishonchli emas. Faqat chat_id dan foydalaning."
            )
            return {"status": "error", "reason": "chat_id_required"}

        try:
            int(target_chat_id)
        except ValueError:
            await _send_message(chat_id=chat_id, text="❌ chat_id raqam bo'lishi kerak.")
            return {"status": "error", "reason": "chat_id_not_numeric"}

        target_username = _normalize_username(parts[2]) if len(parts) > 2 else None
        await _upsert_recipient(
            session=session,
            chat_id=target_chat_id,
            telegram_username=target_username,
            full_name=None,
            added_by_chat_id=chat_id
        )
        await session.commit()

        await _send_message(chat_id=chat_id, text=f"✅ User qo'shildi: {target_chat_id}")
        return {"status": "success", "command": command}

    if command == "/remove_user":
        if len(parts) < 2:
            await _send_message(
                chat_id=chat_id,
                text="ℹ️ Format: /remove_user <chat_id>\nMisol: /remove_user 123456789"
            )
            return {"status": "error", "reason": "invalid_arguments"}

        target_chat_id = parts[1].strip()
        removed = await _remove_recipient(session, target_chat_id)
        await session.commit()
        if removed:
            await _send_message(chat_id=chat_id, text=f"🗑 User o'chirildi: {target_chat_id}")
        else:
            await _send_message(chat_id=chat_id, text="📭 Bunday active user topilmadi.")
        return {"status": "success", "command": command}

    if command == "/run":
        stats = await process_due_recall_notifications(session)
        await _send_message(
            chat_id=chat_id,
            text=(
                "🚀 Reminder run natijasi\n\n"
                f"👥 recipients: {stats.get('recipients_count', 0)}\n"
                f"🧾 customers: {stats.get('customers_checked', 0)}\n"
                f"🔔 due: {stats.get('due_notifications', 0)}\n"
                f"✅ sent: {stats.get('sent', 0)}\n"
                f"⚠️ failed: {stats.get('failed', 0)}\n"
                f"🛡 skipped duplicate: {stats.get('skipped_duplicate', 0)}"
            )
        )
        return {"status": "success", "command": command, "stats": stats}

    if command == "/run_stats":
        await _send_message(
            chat_id=chat_id,
            text=(
                "⛔ 3 kunlik statistikani hammaga yuborish o'chirilgan.\n"
                "Kerak bo'lsa /stats_menu orqali individual statistika oling."
            )
        )
        return {"status": "disabled", "command": command}

    await _send_message(chat_id=chat_id, text=_admin_help_text(), reply_markup=_admin_keyboard())
    return {"status": "ignored", "reason": "unknown_command"}


@router.post("/telegram-webhook", summary="Recall bot webhook")
async def recall_bot_webhook(
    payload: TelegramWebhookPayload,
    session: AsyncSession = Depends(get_async_session)
):
    if not bot:
        return {"status": "error", "reason": "TELEGRAM_RECALL_BOT_TOKEN sozlanmagan"}

    message = payload.message or payload.edited_message
    if not message or not message.text:
        return {"status": "ignored", "reason": "no_text"}

    return await _handle_command(message, session)
