"""
Admin Statistics and AI Summary Generator
Har bir user uchun statistika va AI xulosasini yaratadi
Excel file yaratish va Telegram ga yuborish
"""
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, desc
from datetime import datetime, date, timedelta
from typing import List, Dict, Optional
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from models.admin_models import daily_update_log
from models.user_models import user


def get_working_days_in_month(year: int, month: int) -> tuple[int, List[date]]:
    """
    Oyda necha ish kuni bor (yakshanba kunlarini hisobga olmasdan)

    Returns:
        tuple: (ish_kunlari_soni, yakshanba_kunlari_royxati)
    """
    # Oydagi barcha kunlarni olish
    num_days = calendar.monthrange(year, month)[1]

    working_days = 0
    sundays = []

    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()  # 0=Dushanba, 6=Yakshanba

        if weekday == 6:  # Yakshanba
            sundays.append(current_date)
        else:
            working_days += 1

    return working_days, sundays


async def generate_admin_statistics(
    session: AsyncSession,
    month: Optional[int] = None,
    year: Optional[int] = None
) -> tuple[str, Optional[bytes]]:
    """
    Barcha userlar uchun statistika va AI xulosasi yaratadi
    Oy va yil bo'yicha filtr qiladi, yakshanba kunlarini hisobga olmaydi

    Args:
        session: Database session
        month: Oy (1-12), None bo'lsa hozirgi oy
        year: Yil (masalan 2026), None bo'lsa hozirgi yil

    Returns:
        tuple: (telegram_message, excel_file_bytes)
    """
    # Hozirgi sana
    today = date.today()

    # Oy va yilni aniqlash
    if month is None:
        month = today.month
    if year is None:
        year = today.year

    # Oyda necha ish kuni bor
    working_days, sundays = get_working_days_in_month(year, month)

    # Oy boshlanishi va oxiri
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    # Barcha active userlarni olish
    result = await session.execute(
        select(user.c.id, user.c.name, user.c.surname, user.c.telegram_id)
        .where(user.c.is_active == True)
        .order_by(user.c.name)
    )
    users = result.fetchall()

    if not users:
        return ("❌ Faol foydalanuvchilar topilmadi.", None)

    # Statistika yaratish
    stats_list = []

    for u in users:
        # User uchun bu oydagi yangilanishlar
        updates_result = await session.execute(
            select(daily_update_log)
            .where(
                and_(
                    daily_update_log.c.user_id == u.id,
                    daily_update_log.c.update_date >= first_day,
                    daily_update_log.c.update_date <= last_day
                )
            )
            .order_by(desc(daily_update_log.c.update_date))
        )
        updates = updates_result.fetchall()

        # Yakshanba kunlaridagi updatelarni hisobga olmaymiz
        valid_updates = [
            upd for upd in updates
            if upd.update_date.weekday() != 6  # 6 = Yakshanba
        ]

        # Statistika
        total_days = working_days  # Faqat ish kunlari
        update_days = len(valid_updates)
        update_percentage = round((update_days / total_days) * 100, 1) if total_days > 0 else 0

        # Oxirgi update
        last_update_date = updates[0].update_date if updates else None
        last_update_content = updates[0].update_content if updates else None

        # Days since last update
        days_since_last = None
        if last_update_date:
            days_since_last = (today - last_update_date).days

        # User info
        full_name = f"{u.name} {u.surname}"
        telegram_username = f"#{u.telegram_id}" if u.telegram_id else "#N/A"  # # bilan

        # AI Summary yaratish
        ai_summary = generate_ai_summary(
            full_name=full_name,
            update_days=update_days,
            total_days=total_days,
            update_percentage=update_percentage,
            last_update_content=last_update_content,
            days_since_last=days_since_last
        )

        stats_list.append({
            'name': full_name,
            'username': telegram_username,
            'update_days': update_days,
            'total_days': total_days,
            'percentage': update_percentage,
            'last_update': last_update_date,
            'days_since_last': days_since_last,
            'ai_summary': ai_summary,
            'user_id': u.id
        })

    # Formatlangan xabar yaratish
    report_date = date(year, month, 1)
    message = format_admin_report(stats_list, report_date, working_days, len(sundays))

    # Excel file yaratish
    excel_bytes = generate_excel_report(stats_list, year, month, working_days, len(sundays))

    return message, excel_bytes


def generate_ai_summary(
    full_name: str,
    update_days: int,
    total_days: int,
    update_percentage: float,
    last_update_content: Optional[str],
    days_since_last: Optional[int]
) -> str:
    """
    User uchun AI xulosasi yaratadi (o'zbek tilida)

    Bu versiyada oddiy template-based summary.
    Keyinchalik OpenAI/Claude API bilan boyitish mumkin.
    """
    # Baho berish
    if update_percentage >= 90:
        grade = "A'LO"
        emoji = "🌟"
        comment = "Juda yaxshi natija! Doimiy ravishda update bermoqda."
    elif update_percentage >= 75:
        grade = "YAXSHI"
        emoji = "✅"
        comment = "Yaxshi natija! Izchil ishlayapti."
    elif update_percentage >= 50:
        grade = "O'RTACHA"
        emoji = "⚠️"
        comment = "O'rtacha natija. Ko'proq update berish kerak."
    elif update_percentage >= 25:
        grade = "PAST"
        emoji = "❌"
        comment = "Past ko'rsatkich. Update berish kerak!"
    else:
        grade = "JUDA PAST"
        emoji = "🚨"
        comment = "Juda kam update! Darhol e'tibor berish kerak."

    # Oxirgi update haqida
    if days_since_last is not None:
        if days_since_last == 0:
            last_update_info = "Bugun update bergan"
        elif days_since_last == 1:
            last_update_info = "Kecha update bergan"
        elif days_since_last <= 3:
            last_update_info = f"{days_since_last} kun oldin update bergan"
        elif days_since_last <= 7:
            last_update_info = f"{days_since_last} kun oldin update bergan (e'tibor bering!)"
        else:
            last_update_info = f"{days_since_last} kun oldin update bergan (juda uzoq vaqt!)"
    else:
        last_update_info = "Hech qachon update bermagan"

    # Trend tahlili
    if update_percentage >= 75:
        trend = "Faol va izchil ishlamoqda"
    elif update_percentage >= 50:
        trend = "Yaxshiroq natija uchun izchillikni oshirish kerak"
    elif update_percentage >= 25:
        trend = "Update berish chastotasini sezilarli oshirish talab qilinadi"
    else:
        trend = "Darhol rahbariyat e'tiboriga havola qilish kerak"

    # Tavsiya
    if update_percentage >= 90:
        recommendation = "Davom eting! Ajoyib natija."
    elif update_percentage >= 75:
        recommendation = "Yaxshi ish! 100% ga intilish mumkin."
    elif update_percentage >= 50:
        recommendation = "Har kuni update berishga harakat qiling."
    elif update_percentage >= 25:
        recommendation = "Update berish rejasini tuzish tavsiya etiladi."
    else:
        recommendation = "Zudlik bilan rahbariyat bilan gaplashish kerak."

    # Summary yig'ish
    summary = f"""
{emoji} *BAHO: {grade}* ({update_percentage}%)

📊 *Statistika:*
   • {update_days}/{total_days} kun update bergan
   • {last_update_info}

💡 *Tahlil:*
   {comment}
   {trend}

📝 *Tavsiya:*
   {recommendation}
"""

    return summary.strip()


def format_admin_report(stats_list: List[Dict], report_date: date, working_days: int, sundays_count: int) -> str:
    """
    Admin uchun to'liq hisobotni formatlaydi
    """
    month_name_uz = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
        9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
    }

    month_str = f"{month_name_uz[report_date.month]} {report_date.year}"

    # Sarlavha
    message = f"""
╔════════════════════════════════════════════╗
║         📊 ADMIN STATISTIKA HISOBOTI        ║
╚════════════════════════════════════════════╝

📅 *Davr:* {month_str}
📆 *Ish kunlari:* {working_days} kun (yakshanba: {sundays_count} kun)
👥 *Xodimlar soni:* {len(stats_list)}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""

    # Har bir user uchun statistika
    for i, stat in enumerate(stats_list, 1):
        user_block = f"""
*{i}. {stat['name']}* ({stat['username']})

📈 *Update foizi:* {stat['percentage']}% ({stat['update_days']}/{stat['total_days']} kun)

{stat['ai_summary']}

{'━' * 44}
"""
        message += user_block

    # Umumiy xulosa
    avg_percentage = sum(s['percentage'] for s in stats_list) / len(stats_list)
    total_updates = sum(s['update_days'] for s in stats_list)

    # Top 3 performerlar
    top_3 = sorted(stats_list, key=lambda x: x['percentage'], reverse=True)[:3]
    bottom_3 = sorted(stats_list, key=lambda x: x['percentage'])[:3]

    message += f"""

╔════════════════════════════════════════════╗
║              🎯 UMUMIY XULOSA              ║
╚════════════════════════════════════════════╝

📊 *O'rtacha ko'rsatkich:* {round(avg_percentage, 1)}%
📝 *Jami updatelar:* {total_updates}
📈 *Kun boshiga o'rtacha:* {round(total_updates / len(stats_list), 1)} kun

🌟 *TOP 3 Faol Xodimlar:*
"""

    for i, stat in enumerate(top_3, 1):
        emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉"
        message += f"   {emoji} {stat['name']} - {stat['percentage']}%\n"

    message += f"""
⚠️ *Eng Kam Faol 3 Xodim:*
"""

    for i, stat in enumerate(bottom_3, 1):
        message += f"   {i}. {stat['name']} - {stat['percentage']}%\n"

    message += """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *Tavsiya:*
Kam faol xodimlar bilan individual suhbat o'tkazish tavsiya etiladi.

📞 *Aloqa:* admin@company.uz
"""

    return message


def generate_excel_report(
    stats_list: List[Dict],
    year: int,
    month: int,
    working_days: int,
    sundays_count: int
) -> bytes:
    """
    Excel hisobotini yaratadi

    Returns:
        bytes: Excel file (xlsx) bytes
    """
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month:02d}.{year}"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header qo'shish
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"📊 ADMIN STATISTIKA HISOBOTI - {month:02d}.{year}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = cell_alignment
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)

    # Info qo'shish
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f"📆 Ish kunlari: {working_days} kun | Yakshanba: {sundays_count} kun | Xodimlar: {len(stats_list)}"
    info_cell.alignment = cell_alignment

    # Ustun nomlari
    headers = ['№', 'Ism Familiya', 'Username', 'Update Kunlari', 'Ish Kunlari', 'Foiz (%)', 'Baho', 'Tavsiya']
    ws.append(headers)

    # Header stilini qo'llash
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        cell.border = border

    # Ma'lumotlarni qo'shish
    for idx, stat in enumerate(stats_list, 1):
        # Baho aniqlash
        percentage = stat['percentage']
        if percentage >= 90:
            grade = "A'LO 🌟"
            recommendation = "Davom eting!"
        elif percentage >= 75:
            grade = "YAXSHI ✅"
            recommendation = "Yaxshi ish!"
        elif percentage >= 50:
            grade = "O'RTACHA ⚠️"
            recommendation = "Yaxshilash kerak"
        elif percentage >= 25:
            grade = "PAST ❌"
            recommendation = "E'tibor bering!"
        else:
            grade = "JUDA PAST 🚨"
            recommendation = "Zudlik bilan choralar!"

        row_data = [
            idx,
            stat['name'],
            stat['username'],
            stat['update_days'],
            stat['total_days'],
            stat['percentage'],
            grade,
            recommendation
        ]

        ws.append(row_data)

        # Cell stilini qo'llash
        row_num = idx + 3
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = border

            # Foiz bo'yicha rang berish
            if col_num == 6:  # Foiz ustuni
                if percentage >= 90:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif percentage >= 75:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif percentage >= 50:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

    # Ustun kengliklarini sozlash
    column_widths = [5, 25, 15, 15, 12, 10, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Umumiy statistika qo'shish
    summary_row = len(stats_list) + 5
    ws.merge_cells(f'A{summary_row}:H{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "🎯 UMUMIY XULOSA"
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.alignment = cell_alignment
    summary_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_cell.font = Font(bold=True, color="FFFFFF", size=12)

    # O'rtacha foiz
    avg_percentage = sum(s['percentage'] for s in stats_list) / len(stats_list)
    total_updates = sum(s['update_days'] for s in stats_list)

    summary_row += 1
    ws[f'A{summary_row}'] = "O'rtacha foiz:"
    ws[f'B{summary_row}'] = f"{round(avg_percentage, 1)}%"
    ws[f'C{summary_row}'] = "Jami updatelar:"
    ws[f'D{summary_row}'] = total_updates

    # TOP 3
    summary_row += 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    ws[f'A{summary_row}'] = "🌟 TOP 3 FAOL XODIMLAR"
    ws[f'A{summary_row}'].font = Font(bold=True)

    top_3 = sorted(stats_list, key=lambda x: x['percentage'], reverse=True)[:3]
    for i, stat in enumerate(top_3, 1):
        summary_row += 1
        emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉"
        ws[f'A{summary_row}'] = f"{emoji} {i}"
        ws[f'B{summary_row}'] = stat['name']
        ws[f'C{summary_row}'] = f"{stat['percentage']}%"

    # BOTTOM 3
    summary_row += 2
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    ws[f'A{summary_row}'] = "⚠️ ENG KAM FAOL 3 XODIM"
    ws[f'A{summary_row}'].font = Font(bold=True)

    bottom_3 = sorted(stats_list, key=lambda x: x['percentage'])[:3]
    for i, stat in enumerate(bottom_3, 1):
        summary_row += 1
        ws[f'A{summary_row}'] = i
        ws[f'B{summary_row}'] = stat['name']
        ws[f'C{summary_row}'] = f"{stat['percentage']}%"

    # Excel file ni bytes ga aylantirish
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.read()


async def generate_user_daily_report(
    session: AsyncSession,
    user_id: int,
    month: int,
    year: int
) -> Optional[bytes]:
    """
    Generate daily attendance report for a specific user
    Shows which days the user submitted updates

    Args:
        session: Database session
        user_id: User ID
        month: Month (1-12)
        year: Year

    Returns:
        bytes: Excel file bytes with daily attendance calendar
    """
    # Get user info
    user_result = await session.execute(
        select(user.c.id, user.c.name, user.c.surname, user.c.telegram_id)
        .where(user.c.id == user_id)
    )
    user_data = user_result.fetchone()

    if not user_data:
        return None

    full_name = f"{user_data.name} {user_data.surname}"
    telegram_username = f"#{user_data.telegram_id}" if user_data.telegram_id else "#N/A"

    # Get working days
    working_days, sundays = get_working_days_in_month(year, month)

    # Get all days in month
    num_days = calendar.monthrange(year, month)[1]
    first_day = date(year, month, 1)
    last_day = date(year, month, num_days)

    # Get user's updates for this month
    updates_result = await session.execute(
        select(daily_update_log.c.update_date, daily_update_log.c.update_content, daily_update_log.c.is_valid)
        .where(
            and_(
                daily_update_log.c.user_id == user_id,
                daily_update_log.c.update_date >= first_day,
                daily_update_log.c.update_date <= last_day
            )
        )
        .order_by(daily_update_log.c.update_date)
    )
    updates = {row.update_date: {"content": row.update_content, "valid": row.is_valid} for row in updates_result.fetchall()}

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month:02d}.{year}"

    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    sunday_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    absent_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"📊 {full_name} ({telegram_username}) - Kunlik hisoboti"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = cell_alignment
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)

    # Info row
    month_names_uz = {
        1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
        5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
        9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
    }
    month_name = month_names_uz[month]

    ws.merge_cells('A2:F2')
    info_cell = ws['A2']
    info_cell.value = f"📅 {month_name} {year} | Ish kunlari: {working_days} | Yakshanba: {len(sundays)}"
    info_cell.alignment = cell_alignment

    # Column headers
    headers = ['Kun', 'Sana', 'Hafta kuni', 'Status', 'Update', 'Izoh']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_alignment
        cell.border = border

    # Day names in Uzbek
    weekday_names = {
        0: "Dushanba",
        1: "Seshanba",
        2: "Chorshanba",
        3: "Payshanba",
        4: "Juma",
        5: "Shanba",
        6: "Yakshanba"
    }

    # Add daily data
    update_count = 0
    for day in range(1, num_days + 1):
        current_date = date(year, month, day)
        weekday = current_date.weekday()
        weekday_name = weekday_names[weekday]

        # Check if user submitted update
        has_update = current_date in updates

        if weekday == 6:  # Sunday
            status = "🏖️ Dam olish"
            update_text = "-"
            comment = "Yakshanba"
            fill_color = sunday_fill
        elif has_update:
            status = "✅ Topshirgan"
            update_data = updates[current_date]
            # Truncate update content for display
            update_text = update_data["content"][:50] + "..." if len(update_data["content"]) > 50 else update_data["content"]
            comment = "Valid" if update_data["valid"] else "Invalid"
            fill_color = present_fill
            update_count += 1
        else:
            status = "❌ Topshirmagan"
            update_text = "-"
            comment = "Yo'q"
            fill_color = absent_fill

        row_data = [
            day,
            current_date.strftime("%d.%m.%Y"),
            weekday_name,
            status,
            update_text,
            comment
        ]

        ws.append(row_data)

        # Apply styles
        row_num = day + 3
        for col_num in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = border

            # Apply color based on status
            if col_num >= 4:  # Status and onwards
                cell.fill = fill_color

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 12

    # Summary section
    summary_row = num_days + 5
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = "📊 XULOSA"
    summary_cell.font = Font(bold=True, size=12)
    summary_cell.alignment = cell_alignment
    summary_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_cell.font = Font(bold=True, color="FFFFFF", size=12)

    # Calculate percentage
    percentage = round((update_count / working_days) * 100, 1) if working_days > 0 else 0

    summary_row += 1
    ws[f'A{summary_row}'] = "Jami ish kunlari:"
    ws[f'B{summary_row}'] = working_days
    ws[f'C{summary_row}'] = "Update bergan:"
    ws[f'D{summary_row}'] = update_count

    summary_row += 1
    ws[f'A{summary_row}'] = "Yakshanba kunlari:"
    ws[f'B{summary_row}'] = len(sundays)
    ws[f'C{summary_row}'] = "Foiz:"
    ws[f'D{summary_row}'] = f"{percentage}%"
    ws[f'D{summary_row}'].font = Font(bold=True, size=14)

    # Convert to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.read()
